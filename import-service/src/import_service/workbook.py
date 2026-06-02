from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .models import ContributionFlowRow, FranchiseMonthRow, OverviewCheck, SheetProfile, SiteMonthRow, WorkbookInspection


SHEET_RULES: dict[str, dict[str, int | None]] = {
    "总表-加盟商": {"header_start_row": 1, "header_end_row": 3, "data_start_row": 5, "total_row": 4},
    "总表-网点": {"header_start_row": 1, "header_end_row": 3, "data_start_row": 5, "total_row": 4},
    "总表-一口价": {"header_start_row": 2, "header_end_row": 2, "data_start_row": 3, "total_row": 1},
    "辽宁区域贡献": {"header_start_row": 1, "header_end_row": 2, "data_start_row": 3, "total_row": None},
    "加盟商贡献": {"header_start_row": 1, "header_end_row": 2, "data_start_row": 3, "total_row": None},
    "出港考核、派费补贴": {"header_start_row": 1, "header_end_row": 1, "data_start_row": 2, "total_row": None},
    "包仓费明细": {"header_start_row": 1, "header_end_row": 1, "data_start_row": 2, "total_row": None},
    "运营管理类汇总表": {"header_start_row": 1, "header_end_row": 1, "data_start_row": 2, "total_row": None},
}

WEIGHT_BANDS = ["0.3", "0.5", "1", "2", "3.2", "4", "5.2", "6", "7", "8", "9", "10.3", "＞10.3"]

CONTRIBUTION_GROUP_STARTS = {
    "ticket_count": 6,
    "ticket_share": 20,
    "weight_total": 34,
    "four_fee_total": 48,
    "settlement_price": 62,
    "dispatch_fee": 76,
    "contribution_total": 90,
    "unit_four_fee": 104,
    "unit_settlement_price": 118,
    "unit_dispatch_fee": 132,
    "unit_contribution": 146,
    "kg_contribution": 160,
}


def _num(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _flag_yes_no(value: object) -> bool | None:
    raw = _text(value)
    if raw in {"是", "Y", "Yes", "true", "True", "1"}:
        return True
    if raw in {"否", "N", "No", "false", "False", "0"}:
        return False
    return None


def _cell(row: list[object], index: int) -> object | None:
    return row[index] if index < len(row) else None


def inspect_workbook(path: str | Path) -> WorkbookInspection:
    source = Path(path)
    workbook = load_workbook(source, read_only=True, data_only=True)

    sheets: list[SheetProfile] = []
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        rule = SHEET_RULES.get(sheet_name, {})
        sheets.append(
            SheetProfile(
                name=sheet_name,
                max_row=ws.max_row or 0,
                max_col=ws.max_column or 0,
                header_start_row=rule.get("header_start_row"),
                header_end_row=rule.get("header_end_row"),
                data_start_row=rule.get("data_start_row"),
                total_row=rule.get("total_row"),
            )
        )

    overview = extract_overview_check(workbook)
    return WorkbookInspection(
        path=str(source),
        sheet_count=len(workbook.sheetnames),
        sheets=sheets,
        overview=overview,
    )


def infer_period_month(workbook) -> str:
    if "总表-加盟商" in workbook.sheetnames:
        ws = workbook["总表-加盟商"]
        first_data_row = next(ws.iter_rows(min_row=5, max_row=5, values_only=True), None)
        if first_data_row:
            value = _text(_cell(list(first_data_row), 3))
            if value:
                return value
    return ""


def parse_franchise_month_rows(path: str | Path) -> list[FranchiseMonthRow]:
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    if "总表-加盟商" not in workbook.sheetnames:
        return []

    ws = workbook["总表-加盟商"]
    rows: list[FranchiseMonthRow] = []

    for raw in ws.iter_rows(min_row=5, values_only=True):
        row = list(raw)
        franchise_name = _text(_cell(row, 4))
        if not franchise_name:
            continue

        rows.append(
            FranchiseMonthRow(
                period_month=_text(_cell(row, 3)),
                franchise_name=franchise_name,
                daily_over_5000_flag=_flag_yes_no(_cell(row, 1)),
                outbound_tickets=_num(_cell(row, 6)),
                outbound_weight=_num(_cell(row, 7)),
                outbound_avg_weight=_num(_cell(row, 8)),
                waybill_fee=_num(_cell(row, 9)),
                transfer_fee=_num(_cell(row, 10)),
                warehouse_fee=_num(_cell(row, 11)),
                operation_fee=_num(_cell(row, 12)),
                dispatch_fee=_num(_cell(row, 14)),
                one_price_rebate=_num(_cell(row, 15)),
                outbound_contribution=_num(_cell(row, 35)),
                outbound_unit_contribution=_num(_cell(row, 36)),
                outbound_kg_contribution=_num(_cell(row, 37)),
                inbound_signed_tickets=_num(_cell(row, 38)),
                inbound_weight=_num(_cell(row, 39)),
                inbound_dispatch_income=_num(_cell(row, 41)),
                inbound_dispatch_cost=_num(_cell(row, 44)),
                deduction_total=_num(_cell(row, 66)),
                inbound_contribution=_num(_cell(row, 68)),
                total_contribution=_num(_cell(row, 69)),
                outbound_pass_contribution=_num(_cell(row, 70)),
                inbound_pass_contribution=_num(_cell(row, 71)),
            )
        )

    return rows


def parse_contribution_flow_rows(
    path: str | Path,
    *,
    scope_type: str,
    include_total_rows: bool = False,
    region_code: str = "LN",
) -> list[ContributionFlowRow]:
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    if scope_type == "region":
        sheet_name = "辽宁区域贡献"
    elif scope_type == "franchise":
        sheet_name = "加盟商贡献"
    else:
        raise ValueError("scope_type must be 'region' or 'franchise'")

    if sheet_name not in workbook.sheetnames:
        return []

    period_month = infer_period_month(workbook)
    ws = workbook[sheet_name]
    rows: list[ContributionFlowRow] = []

    for raw in ws.iter_rows(min_row=3, values_only=True):
        row = list(raw)
        destination = _text(_cell(row, 5))
        if not destination:
            continue
        if not include_total_rows and destination in {"小计", "合计", "总计"}:
            continue

        franchise_name = _text(_cell(row, 1)) if scope_type == "franchise" else None
        if scope_type == "franchise" and not franchise_name:
            continue

        row_region_code = region_code if scope_type == "region" else None

        for offset, weight_band in enumerate(WEIGHT_BANDS):
            rows.append(
                ContributionFlowRow(
                    period_month=period_month,
                    scope_type=scope_type,
                    region_code=row_region_code,
                    franchise_name=franchise_name,
                    destination_province=destination,
                    weight_band=weight_band,
                    ticket_count=_num(_cell(row, CONTRIBUTION_GROUP_STARTS["ticket_count"] + offset)),
                    ticket_share=_num(_cell(row, CONTRIBUTION_GROUP_STARTS["ticket_share"] + offset)),
                    weight_total=_num(_cell(row, CONTRIBUTION_GROUP_STARTS["weight_total"] + offset)),
                    four_fee_total=_num(_cell(row, CONTRIBUTION_GROUP_STARTS["four_fee_total"] + offset)),
                    settlement_price=_num(_cell(row, CONTRIBUTION_GROUP_STARTS["settlement_price"] + offset)),
                    dispatch_fee=_num(_cell(row, CONTRIBUTION_GROUP_STARTS["dispatch_fee"] + offset)),
                    contribution_total=_num(_cell(row, CONTRIBUTION_GROUP_STARTS["contribution_total"] + offset)),
                    unit_four_fee=_num(_cell(row, CONTRIBUTION_GROUP_STARTS["unit_four_fee"] + offset)),
                    unit_settlement_price=_num(_cell(row, CONTRIBUTION_GROUP_STARTS["unit_settlement_price"] + offset)),
                    unit_dispatch_fee=_num(_cell(row, CONTRIBUTION_GROUP_STARTS["unit_dispatch_fee"] + offset)),
                    unit_contribution=_num(_cell(row, CONTRIBUTION_GROUP_STARTS["unit_contribution"] + offset)),
                    kg_contribution=_num(_cell(row, CONTRIBUTION_GROUP_STARTS["kg_contribution"] + offset)),
                )
            )

    return rows


def parse_site_month_rows(path: str | Path) -> list[SiteMonthRow]:
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    if "总表-网点" not in workbook.sheetnames:
        return []

    ws = workbook["总表-网点"]
    rows: list[SiteMonthRow] = []

    for raw in ws.iter_rows(min_row=5, values_only=True):
        row = list(raw)
        franchise_name = _text(_cell(row, 4))
        site_name = _text(_cell(row, 5))
        if not franchise_name or not site_name:
            continue

        rows.append(
            SiteMonthRow(
                period_month=_text(_cell(row, 3)),
                franchise_name=franchise_name,
                site_name=site_name,
                site_status=_text(_cell(row, 0)) or None,
                daily_over_5000_flag=_flag_yes_no(_cell(row, 1)),
                outbound_tickets=_num(_cell(row, 6)),
                outbound_weight=_num(_cell(row, 7)),
                outbound_contribution=_num(_cell(row, 35)),
                inbound_signed_tickets=_num(_cell(row, 38)),
                inbound_contribution=_num(_cell(row, 68)),
                deduction_total=_num(_cell(row, 66)),
                total_contribution=_num(_cell(row, 69)),
            )
        )

    return rows


def extract_overview_check(workbook) -> OverviewCheck:
    franchise_total = None
    site_total = None

    if "总表-加盟商" in workbook.sheetnames:
        ws = workbook["总表-加盟商"]
        franchise_total = next(ws.iter_rows(min_row=4, max_row=4, values_only=True), None)

    if "总表-网点" in workbook.sheetnames:
        ws = workbook["总表-网点"]
        site_total = next(ws.iter_rows(min_row=4, max_row=4, values_only=True), None)

    return OverviewCheck(
        franchise_count=_num(franchise_total[4]) if franchise_total else None,
        site_count=_num(site_total[5]) if site_total else None,
        outbound_tickets=_num(franchise_total[6]) if franchise_total else None,
        outbound_weight=_num(franchise_total[7]) if franchise_total else None,
        inbound_signed_tickets=_num(franchise_total[38]) if franchise_total else None,
        outbound_contribution=_num(franchise_total[35]) if franchise_total else None,
        inbound_contribution=_num(franchise_total[68]) if franchise_total else None,
        total_contribution=_num(franchise_total[69]) if franchise_total else None,
        deduction_total=_num(franchise_total[66]) if franchise_total else None,
    )
